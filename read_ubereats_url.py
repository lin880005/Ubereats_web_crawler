from selenium import webdriver # 瀏覽器
from selenium.webdriver.common.by import By # 選取器
from selenium.webdriver.support.wait import WebDriverWait # 網站等待
from selenium.webdriver.support import expected_conditions as EC # 元素狀態判斷
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
import time
import json
import openpyxl
import re
import pandas as pd

def data_clean(text):
    # 清洗excel中的非法字符，都是不常见的不可显示字符，例如退格，响铃等
    ILLEGAL_CHARACTERS_RE = re.compile(r"[\000-\010]|[\013-\014]|[\016-\037]")
    text = ILLEGAL_CHARACTERS_RE.sub(r"", text)
    return text

options = Options()
options.add_argument("--disable-notifications") # 關閉彈出式視窗
driver = webdriver.Chrome(chrome_options = options)

driver.get("https://www.ubereats.com/tw")

getblock = driver.find_element(By.XPATH, '//*[@placeholder="輸入外送地址"]')
getblock.send_keys('高雄市')
time.sleep(1)
getblock.send_keys('\ue007') # 按下Enter
time.sleep(3)


wb = openpyxl.Workbook()
ws = wb.active



ws["A1"] = "餐廳名稱"
ws["B1"] = "餐廳類型"
ws["C1"] = "餐廳總評分"
ws["D1"] = "地址"
ws["E1"] = "經度"
ws["F1"] = "緯度"
ws["G1"] = "訂餐網址"


df = pd.read_excel('Uber_eats高雄市餐廳網址.xlsx')

# 提取網址欄位的數據
urls = df['URL']

count = 0

for store in urls:
    driver.get(store)
    time.sleep(3)
    try:         # 👈👀 正常接單的店家，可直接點選到 "詳細資訊"，走try 📌

        detail = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[text()="詳細資訊"]')))
        detail.click()
        time.sleep(5)
        soup = BeautifulSoup(driver.page_source, "html.parser")
        #print(soup)

        info = soup.find_all("main", id="main-content")[0].script.text
        dic_info = json.loads(info)
        #print(info)
        name = dic_info["name"]  # 店名
        
        try:
            sc = dic_info["aggregateRating"]["ratingValue"]  # 總評分
            ad = dic_info["address"]["streetAddress"]                       # 👈👀 先不要抓地址 📌
            type = dic_info["servesCuisine"][0]  # 類型
            lo = dic_info["geo"]["longitude"]#經度
            la = dic_info["geo"]["latitude"]#緯度
        except:
            sc = "NoRating"
            ad = ""
            type = ""
            lo = "" #經度
            la = "" #緯度
        count+=1
        print(f"====第{count}間====")

    except:       # 👈👀 已不接單的店家，有彈出視窗導致無法直接點選到 "詳細資訊"，走except 📌

        closeButton = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'button[aria-label="關閉"]')))
        closeButton.click()
        
        # ☝👀 多一個步驟，要先把彈出的視窗關閉 📌
        # 👇👀 後面一大坨就跟前面的try一樣，點選 "詳細資訊" 抓店家資料 📌
        
        detail = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[text()="詳細資訊"]')))
        detail.click()
        time.sleep(5)
        soup = BeautifulSoup(driver.page_source, "html.parser")
        #print(soup)

        info = soup.find_all("main", id="main-content")[0].script.text
        dic_info = json.loads(info)
        #print(info)
        name = dic_info["name"]  # 店名
        
        try:
            sc = dic_info["aggregateRating"]["ratingValue"]  # 總評分
            ad = dic_info["address"]["streetAddress"]    # 👈👀 先不要抓地址 📌
            type = dic_info["servesCuisine"][0]  # 類型
            lo = dic_info["geo"]["longitude"]#經度
            la = dic_info["geo"]["latitude"]#緯度                   
        except:
            sc = "NoRating"
            ad = ""
            type = ""
            lo = "" #經度
            la = "" #緯度
        count+=1
        print(f"====第{count}間====")
 

    ws.append([data_clean(name),type,sc,ad,lo,la, store])
    wb.save("Uber_eats高雄市.xlsx")
driver.quit()